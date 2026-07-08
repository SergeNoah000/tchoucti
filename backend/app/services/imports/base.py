"""Socle générique d'import par templates Excel (.xlsx).

Chaque entité importable fournit un ``Importer`` : la liste de ses colonnes
(avec aide + exemple + valeurs autorisées), une validation par ligne et une
création par ligne. Le socle génère le template .xlsx, parse un fichier
re-uploadé, et orchestre l'aperçu (dry-run) puis la création réelle.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.ext.asyncio import AsyncSession


@dataclass(frozen=True)
class Choice:
    """Valeur autorisée d'une colonne (liste déroulante)."""

    value: str          # valeur canonique stockée
    label: str          # libellé affiché (FR) dans la liste déroulante


@dataclass(frozen=True)
class ImportColumn:
    key: str                              # clé interne (champ normalisé)
    header: str                           # en-tête affiché dans le template
    required: bool = False
    help: str = ""                        # note explicative (commentaire cellule)
    example: str = ""                     # exemple de valeur
    choices: tuple[Choice, ...] = ()      # si renseigné → liste déroulante

    def normalize(self, raw: Any) -> Optional[str]:
        """Nettoie une valeur de cellule en chaîne (ou None si vide)."""
        if raw is None:
            return None
        s = str(raw).strip()
        if s == "":
            return None
        if self.choices:
            low = s.lower()
            for c in self.choices:
                if low == c.value.lower() or low == c.label.lower():
                    return c.value
            # valeur hors liste : renvoyée telle quelle, la validation lèvera l'erreur
        return s


@dataclass
class RowResult:
    index: int                            # n° de ligne (1 = 1re ligne de données)
    values: dict[str, Any]                # valeurs normalisées (clé interne → valeur)
    errors: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors


class Importer:
    """Base d'un importer d'entité. À sous-classer."""

    entity: str = ""              # clé technique (ex. "members")
    label: str = ""               # libellé (ex. "Membres")
    description: str = ""         # phrase d'aide
    sheet_title: str = "Données"
    sheet_key: str = ""           # clé de feuille (défaut : entity)
    columns: list[ImportColumn] = []

    @property
    def sheets(self) -> list["Importer"]:
        """Feuilles composant ce classeur. Un importer simple = une feuille."""
        return [self]

    @property
    def key(self) -> str:
        return self.sheet_key or self.entity

    # ── À implémenter par les sous-classes ───────────────────────────────────
    async def validate_row(
        self, db: AsyncSession, association_id, values: dict[str, Any], ctx: dict
    ) -> tuple[Optional[dict], list[str]]:
        """Valide une ligne ; renvoie (payload prêt à créer, liste d'erreurs).

        ``ctx`` est partagé entre les lignes (caches, doublons intra-fichier).
        """
        raise NotImplementedError

    async def create_row(
        self, db: AsyncSession, association_id, payload: dict, ctx: dict
    ) -> None:
        """Crée l'entité depuis un payload validé (sans commit)."""
        raise NotImplementedError

    async def new_ctx(self, db: AsyncSession, association_id) -> dict:
        """Contexte partagé optionnel (préchargements, caches)."""
        return {}

    async def after_commit(self, db: AsyncSession, ctx: dict) -> None:
        """Hook exécuté APRÈS le commit global (envoi de mails, etc.)."""
        return None

    async def preview_register(self, payload: dict, ctx: dict) -> None:
        """APERÇU uniquement : enregistre dans les caches partagés ce que cette
        ligne CRÉERAIT (clé de liaison), pour que les feuilles aval valident
        comme si l'amont existait — sans rien écrire en base. À surcharger pour
        les feuilles « config » référencées par d'autres feuilles."""
        return None

    # ── Template .xlsx ───────────────────────────────────────────────────────
    def build_template(self) -> bytes:
        """Classeur : un onglet par feuille (self.sheets)."""
        wb = Workbook()
        wb.remove(wb.active)
        for sheet in self.sheets:
            ws = wb.create_sheet(title=sheet.sheet_title[:31])
            sheet._render_template_ws(ws)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _render_template_ws(self, ws) -> None:
        """Remplit un onglet avec les colonnes de CETTE feuille."""
        header_fill = PatternFill("solid", fgColor="0F766E")
        req_fill = PatternFill("solid", fgColor="B45309")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, col in enumerate(self.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col.header)
            cell.font = header_font
            cell.fill = req_fill if col.required else header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            note = col.help
            if col.required:
                note = "OBLIGATOIRE. " + note
            if col.example:
                note = (note + f"\nExemple : {col.example}").strip()
            if col.choices:
                opts = ", ".join(c.label for c in col.choices)
                note = (note + f"\nValeurs : {opts}").strip()
            if note:
                cell.comment = Comment(note, "Tchoucti")
            ws.column_dimensions[get_column_letter(col_idx)].width = max(16, len(col.header) + 4)

            # Liste déroulante sur les lignes de données.
            if col.choices:
                letter = get_column_letter(col_idx)
                dv = DataValidation(
                    type="list",
                    formula1='"' + ",".join(c.label for c in col.choices) + '"',
                    allow_blank=not col.required,
                )
                ws.add_data_validation(dv)
                dv.add(f"{letter}2:{letter}1000")

        # Ligne d'exemple (grisée, italique) — à supprimer avant import.
        example_font = Font(italic=True, color="9CA3AF")
        if any(c.example for c in self.columns):
            for col_idx, col in enumerate(self.columns, start=1):
                ex = col.example
                if col.choices and ex:
                    # affiche le libellé correspondant si l'exemple est une valeur
                    ex = next((c.label for c in col.choices if c.value == ex or c.label == ex), ex)
                cell = ws.cell(row=2, column=col_idx, value=ex or None)
                cell.font = example_font

        ws.freeze_panes = "A2"

    # ── Export .xlsx (aller-retour : même format que l'import) ───────────────
    async def export_rows(
        self, db: AsyncSession, association_id, ctx: dict
    ) -> list[dict[str, Any]]:
        """Renvoie les lignes de CETTE feuille à exporter (clé interne → valeur
        native : str/int/date/bool/enum). À surcharger. Défaut : rien."""
        return []

    def _export_value(self, col: "ImportColumn", value: Any) -> Any:
        """Convertit une valeur native en valeur de cellule, cohérente avec le
        template (label pour les listes, ISO pour les dates, Oui/Non pour les
        booléens) — de sorte que le fichier exporté soit ré-importable."""
        from datetime import date as _date, datetime as _datetime
        from decimal import Decimal as _Decimal

        if value is None:
            return None
        v = value.value if hasattr(value, "value") else value
        if col.choices:
            for c in col.choices:
                if str(v) == c.value:
                    return c.label
            return v
        if isinstance(value, bool):
            return "Oui" if value else "Non"
        if isinstance(value, _datetime):
            return value.date().isoformat()
        if isinstance(value, _date):
            return value.isoformat()
        if isinstance(v, _Decimal):
            # openpyxl n'écrit pas les Decimal : normalise en int si entier.
            f = float(v)
            return int(f) if f.is_integer() else f
        return v

    def _render_export_ws(self, ws, rows: list[dict[str, Any]]) -> None:
        """En-têtes (même style que le template) + lignes de données."""
        header_fill = PatternFill("solid", fgColor="0F766E")
        req_fill = PatternFill("solid", fgColor="B45309")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, col in enumerate(self.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col.header)
            cell.font = header_font
            cell.fill = req_fill if col.required else header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(16, len(col.header) + 4)
            if col.choices:
                letter = get_column_letter(col_idx)
                dv = DataValidation(
                    type="list",
                    formula1='"' + ",".join(c.label for c in col.choices) + '"',
                    allow_blank=not col.required,
                )
                ws.add_data_validation(dv)
                dv.add(f"{letter}2:{letter}100000")
        for r_idx, row in enumerate(rows, start=2):
            for col_idx, col in enumerate(self.columns, start=1):
                ws.cell(row=r_idx, column=col_idx, value=self._export_value(col, row.get(col.key)))
        ws.freeze_panes = "A2"

    async def build_export(self, db: AsyncSession, association_id) -> bytes:
        """Classeur exporté : un onglet par feuille, rempli des données réelles."""
        wb = Workbook()
        wb.remove(wb.active)
        ctx = await self.new_ctx(db, association_id)
        for sheet in self.sheets:
            ws = wb.create_sheet(title=sheet.sheet_title[:31])
            rows = await sheet.export_rows(db, association_id, ctx)
            sheet._render_export_ws(ws, rows)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Parsing d'un fichier re-uploadé ──────────────────────────────────────
    def parse(self, data: bytes) -> list[dict[str, Any]]:
        """Lit l'onglet de CETTE feuille (par titre, sinon le 1er) et renvoie
        ses lignes normalisées (ignore la ligne d'exemple/les lignes vides)."""
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[self.sheet_title] if self.sheet_title in wb.sheetnames else wb.active
        return self._parse_ws(ws)

    def _parse_ws(self, ws) -> list[dict[str, Any]]:
        if ws is None:
            return []
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        # header affiché → colonne
        by_header = {col.header: col for col in self.columns}
        col_map: dict[int, ImportColumn] = {}
        for i, h in enumerate(headers):
            if h in by_header:
                col_map[i] = by_header[h]

        out: list[dict[str, Any]] = []
        for raw_row in rows[1:]:
            values: dict[str, Any] = {}
            non_empty = False
            for i, col in col_map.items():
                raw = raw_row[i] if i < len(raw_row) else None
                norm = col.normalize(raw)
                values[col.key] = norm
                if norm is not None:
                    non_empty = True
            if non_empty:
                out.append(values)
        return out

    def _is_example_row(self, values: dict[str, Any]) -> bool:
        """La ligne d'exemple du template (toutes les valeurs == exemples)."""
        ex = {c.key: c.normalize(c.example) for c in self.columns if c.example}
        if not ex:
            return False
        return all(values.get(k) == v for k, v in ex.items())


class DomainImporter(Importer):
    """Classeur multi-feuilles (un domaine : Caisses, Tontines, Prêts, Aides).

    `sheet_importers` est la liste ORDONNÉE des feuilles. Le ctx est PARTAGÉ
    entre toutes les feuilles (caches de liaison : membres par n°, caisses par
    nom, références…) afin que les feuilles « mouvements » résolvent ce que les
    feuilles « config » ont créé.
    """

    sheet_importers: list[Importer] = []

    @property
    def sheets(self) -> list[Importer]:
        return self.sheet_importers

    async def new_ctx(self, db: AsyncSession, association_id) -> dict:
        # Un seul ctx PLAT partagé par toutes les feuilles : chaque feuille y
        # lit/écrit ses caches (assoc, slugs, seen_*, et les caches de liaison
        # membership_by_number, caisse_by_name… alimentés par les create_row).
        ctx: dict = {}
        for sh in self.sheets:
            sub = await sh.new_ctx(db, association_id)
            for k, v in sub.items():
                ctx.setdefault(k, v)
        return ctx
